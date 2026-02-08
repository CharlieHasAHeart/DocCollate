from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from .constants import (
    CELL_MAP_CHECKBOX,
    CELL_MAP_TEXT,
    CELL_MODE_EMBEDDED,
    CELL_MODE_PURE,
    CHECKED_SYMBOL,
    UNCHECKED_SYMBOL,
    WRAP_TEXT_KEYS,
)


def _resolve_merged_cell(ws, coord: str):
    for cell_range in ws.merged_cells.ranges:
        if coord in cell_range:
            return ws.cell(row=cell_range.min_row, column=cell_range.min_col)
    return ws[coord]


def _set_checkbox(ws, cell_coord: str, checked: bool) -> None:
    cell = _resolve_merged_cell(ws, cell_coord)
    target = CHECKED_SYMBOL if checked else UNCHECKED_SYMBOL
    value = cell.value
    if isinstance(value, str) and value:
        cell.value = value.replace("□", target).replace(UNCHECKED_SYMBOL, target).replace(CHECKED_SYMBOL, target)
    else:
        cell.value = target


def _set_text_cell(ws, coord: str, value: str, wrap: bool) -> None:
    cell = _resolve_merged_cell(ws, coord)
    cell.value = value
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")


def generate_excel(template_path: Path, output_path: Path, data: dict) -> None:
    wb = load_workbook(template_path)
    sheets = wb.worksheets

    for key, (sheet_idx, coord) in CELL_MAP_TEXT.items():
        value = str(data.get(key, "") or "")
        _set_text_cell(sheets[sheet_idx], coord, value, key in WRAP_TEXT_KEYS)

    for key, (sheet_idx, coord) in CELL_MAP_CHECKBOX.items():
        value = data.get(key, False)
        checked = bool(value) if not isinstance(value, str) else value.strip().lower() in {"true", "1", "yes", "y", "是"}
        _set_checkbox(sheets[sheet_idx], coord, checked)

    mode_value = str(data.get("assess__product_mode_val", "pure")).lower()
    if not mode_value:
        mode_value = "embedded" if bool(data.get("assess__is_embedded")) else "pure"
    sheet_idx, coord = CELL_MODE_PURE if mode_value != "embedded" else CELL_MODE_EMBEDDED
    _set_checkbox(sheets[sheet_idx], coord, True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
