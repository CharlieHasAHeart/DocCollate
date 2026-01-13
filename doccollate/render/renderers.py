from __future__ import annotations

from pathlib import Path

from docxtpl import DocxTemplate
from docx import Document as DocxWriter
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from ..core.constants import (
    CELL_MAP_CHECKBOX,
    CELL_MAP_TEXT,
    CELL_MODE_EMBEDDED,
    CELL_MODE_PURE,
    CHECKED_SYMBOL,
    UNCHECKED_SYMBOL,
    WRAP_TEXT_KEYS,
)
from .fill_form import docx_replace_text


def fill_func_table(template_path: Path, output_path: Path, data: dict) -> bool:
    if "product__func_list" not in data:
        return False
    raw_items = data.get("product__func_list")
    if not isinstance(raw_items, list):
        return False
    func_list: list[dict[str, str]] = []
    for item in raw_items:
        if isinstance(item, dict):
            name = str(item.get("name") or item.get("一级功能") or "").strip()
            desc = str(item.get("desc") or item.get("功能描述") or "").strip()
        else:
            name = str(item).strip()
            desc = ""
        if not name:
            continue
        func_list.append({"name": name, "desc": desc})
    if not func_list:
        return False
    tpl = DocxTemplate(template_path)
    context = {"software_name": data.get("app__name", ""), "func_list": func_list}
    tpl.render(context)
    tpl.save(output_path)
    return True


def build_holder_context(contact_info: dict) -> dict:
    def pick(*keys: str) -> str:
        for key in keys:
            value = contact_info.get(key)
            if value:
                return str(value)
        return ""

    return {
        "{{holder__name}}": pick("holder__name", "owner", "name"),
        "{{holder__address}}": pick("holder__address", "address"),
        "{{holder__zip_code}}": pick("holder__zip_code", "zip_code", "holder__postcode", "postcode"),
        "{{holder__contact_name}}": pick("holder__contact_name", "contact_name", "contact"),
        "{{holder__contact_mobile}}": pick("holder__contact_mobile", "contact_mobile", "mobile"),
        "{{holder__contact_email}}": pick("holder__contact_email", "contact_email", "email"),
        "{{holder__contact_landline}}": pick("holder__contact_landline", "contact_landline", "phone", "landline"),
        "{{holder__tech_contact_name}}": pick("holder__tech_contact_name", "tech_contact_name"),
        "{{holder__tech_contact_mobile}}": pick("holder__tech_contact_mobile", "tech_contact_mobile"),
    }


def fill_reg_table(template_path: Path, output_path: Path, data: dict, contact_info: dict) -> bool:
    doc = DocxWriter(template_path)
    env_dev_platform = data.get("env__dev_platform") or data.get("env__sw_dev_platform", "")
    env_dev_lang = data.get("env__dev_lang") or data.get("env__language", "")
    env_run_platform = data.get("env__run_platform") or data.get("env__os", "")
    context = {
        "{{app__name}}": data.get("app__name", ""),
        "{{app__short_name}}": data.get("app__short_name", ""),
        "{{app__version}}": data.get("app__version", ""),
        "{{app__product_type_text}}": data.get("app__product_type_text", ""),
        "{{product__app_domain}}": data.get("product__app_domain", ""),
        "{{app__category_assess}}": data.get("app__category_assess", ""),
        "{{product__service_object}}": data.get("product__service_object", ""),
        "{{product__main_functions}}": data.get("product__main_functions", ""),
        "{{product__tech_specs}}": data.get("product__tech_specs", ""),
        "{{env__dev_platform}}": env_dev_platform,
        "{{env__dev_lang}}": env_dev_lang,
        "{{env__run_platform}}": env_run_platform,
        "{{env__hw_dev_platform}}": data.get("env__hw_dev_platform", ""),
        "{{env__sw_dev_platform}}": data.get("env__sw_dev_platform", ""),
        "{{env__memory_req}}": data.get("env__memory_req", ""),
        "{{env__hardware_model}}": data.get("env__hardware_model", ""),
        "{{env__language}}": data.get("env__language", ""),
        "{{env__database}}": data.get("env__database", ""),
        "{{env__os_version}}": data.get("env__os_version", ""),
        "{{env__server_soft}}": data.get("env__server_soft", ""),
        "{{env__client_soft}}": data.get("env__client_soft", ""),
    }
    context.update(build_holder_context(contact_info))
    docx_replace_text(doc, context)
    doc.save(output_path)
    return True


def set_checkbox(ws, cell_coord: str, checked: bool) -> None:
    cell = _resolve_merged_cell(ws, cell_coord)
    target = CHECKED_SYMBOL if checked else UNCHECKED_SYMBOL
    value = cell.value
    if isinstance(value, str) and value:
        cell.value = value.replace("□", target).replace(UNCHECKED_SYMBOL, target).replace(CHECKED_SYMBOL, target)
    else:
        cell.value = target


def _resolve_merged_cell(ws, coord: str):
    for cell_range in ws.merged_cells.ranges:
        if coord in cell_range:
            return ws.cell(row=cell_range.min_row, column=cell_range.min_col)
    return ws[coord]


def set_text_cell(ws, coord: str, value: str, wrap: bool) -> None:
    cell = _resolve_merged_cell(ws, coord)
    cell.value = value
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")


def fill_assessment_excel(template_path: Path, output_path: Path, data: dict) -> bool:
    wb = load_workbook(template_path)
    sheets = wb.worksheets

    for key, (sheet_idx, coord) in CELL_MAP_TEXT.items():
        value = data.get(key, "") or ""
        wrap = key in WRAP_TEXT_KEYS
        set_text_cell(sheets[sheet_idx], coord, value, wrap)

    for key, (sheet_idx, coord) in CELL_MAP_CHECKBOX.items():
        value = str(data.get(key, ""))
        set_checkbox(sheets[sheet_idx], coord, value.strip().lower() in {"true", "yes", "1", "y", "是"})

    mode_value = str(data.get("assess__product_mode_val", "")).lower()
    if not mode_value:
        mode_value = "embedded" if data.get("assess__is_embedded") else "pure"
    sheet_idx, coord = CELL_MODE_PURE if mode_value != "embedded" else CELL_MODE_EMBEDDED
    set_checkbox(sheets[sheet_idx], coord, True)

    wb.save(output_path)
    return True


def fill_env_table(template_path: Path, output_path: Path, data: dict) -> bool:
    doc = DocxWriter(template_path)
    context = {
        "{{env__server_os}}": data.get("env__server_os", ""),
        "{{env__server_soft}}": data.get("env__server_soft", ""),
        "{{env__server_model}}": data.get("env__server_model", ""),
        "{{env__server_config}}": data.get("env__server_config", ""),
        "{{env__server_id}}": data.get("env__server_id", "厂商设备"),
        "{{env__client_os}}": data.get("env__client_os", ""),
        "{{env__client_soft}}": data.get("env__client_soft", ""),
        "{{env__client_model}}": data.get("env__client_model", ""),
        "{{env__client_config}}": data.get("env__client_config", ""),
        "{{env__client_id}}": data.get("env__client_id", "厂商设备"),
    }
    docx_replace_text(doc, context)
    doc.save(output_path)
    return True
