import argparse
import json
import logging
import os
import re
import calendar
from dataclasses import dataclass
from datetime import date, datetime, timedelta
import random
from pathlib import Path
from typing import Iterable

import yaml
from docxtpl import DocxTemplate
from docx import Document as DocxWriter
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from .fill_form import docx_replace_text, generate_document

logging.basicConfig(level=logging.WARNING, format='%(asctime)s - %(levelname)s - %(message)s')
logging.getLogger("httpx").setLevel(logging.WARNING)

CONFIG_DIR = Path.home() / ".doccollate"
DEFAULT_ENV_FILE = Path(".env")
DEFAULT_CONFIG_FILE = CONFIG_DIR / "soft_copyright.yaml"

CHECKED_SYMBOL = "\u2611"
UNCHECKED_SYMBOL = "\u2610"

TEMPLATE_ENV_MAP = {
    "func": "DOCCOLLATE_TEMPLATE_FUNC",
    "reg": "DOCCOLLATE_TEMPLATE_REG",
    "assess": "DOCCOLLATE_TEMPLATE_ASSESS",
    "env": "DOCCOLLATE_TEMPLATE_ENV",
    "copyright": "DOCCOLLATE_TEMPLATE_COPYRIGHT",
}

CELL_MAP_TEXT = {
    "product__service_object": (0, "B2"),
    "product__main_functions": (0, "B3"),
    "product__tech_specs": (0, "B4"),
    "app__product_type_text": (0, "B5"),
    "env__memory_req": (1, "B3"),
    "env__hardware_model": (1, "E3"),
    "env__os": (1, "B8"),
    "env__language": (1, "E8"),
    "env__database": (1, "B9"),
    "env__soft_scale": (1, "E9"),
    "env__os_version": (1, "B10"),
    "env__hw_dev_platform": (1, "C12"),
    "env__sw_dev_platform": (1, "C14"),
    "assess__workload": (2, "C7"),
    "app__category_assess": (2, "B10"),
    "assess__dev_date": (2, "C5"),
    "assess__completion_date": (2, "C6"),
}

CELL_MAP_CHECKBOX = {
    "assess__support_floppy": (1, "A4"),
    "assess__support_sound": (1, "D4"),
    "assess__support_cdrom": (1, "A5"),
    "assess__support_gpu": (1, "D5"),
    "assess__support_other": (1, "A6"),
    "assess__is_self_dev": (2, "A2"),
    "assess__has_docs": (2, "A3"),
    "assess__has_source": (2, "A4"),
}

CELL_MODE_PURE = (2, "C8")
CELL_MODE_EMBEDDED = (2, "C9")

WRAP_TEXT_KEYS = {
    "product__service_object",
    "product__main_functions",
    "product__tech_specs",
    "app__product_type_text",
    "env__os",
    "env__hw_dev_platform",
    "env__sw_dev_platform",
    "app__category_assess",
}


PRODUCT_TYPE_PREFIX = {
    "Operating System": "\u7cfb\u7edf\u8f6f\u4ef6-",
    "Chinese Processing System": "\u7cfb\u7edf\u8f6f\u4ef6-",
    "Network System": "\u7cfb\u7edf\u8f6f\u4ef6-",
    "Embedded Operating System": "\u7cfb\u7edf\u8f6f\u4ef6-",
    "Other(System)": "\u7cfb\u7edf\u8f6f\u4ef6-",
    "Programming Language": "\u652f\u6301\u8f6f\u4ef6-",
    "Database System Design": "\u652f\u6301\u8f6f\u4ef6-",
    "Tools": "\u652f\u6301\u8f6f\u4ef6-",
    "Network Communication Software": "\u652f\u6301\u8f6f\u4ef6-",
    "Middleware": "\u652f\u6301\u8f6f\u4ef6-",
    "Other(Support)": "\u652f\u6301\u8f6f\u4ef6-",
    "Industry Management Software": "\u5e94\u7528\u8f6f\u4ef6-",
    "Office Software": "\u5e94\u7528\u8f6f\u4ef6-",
    "Pattern Recognition Software": "\u5e94\u7528\u8f6f\u4ef6-",
    "Graphics Software": "\u5e94\u7528\u8f6f\u4ef6-",
    "Control Software": "\u5e94\u7528\u8f6f\u4ef6-",
    "Network Application Software": "\u5e94\u7528\u8f6f\u4ef6-",
    "Information Management Software": "\u5e94\u7528\u8f6f\u4ef6-",
    "Database Management Application Software": "\u5e94\u7528\u8f6f\u4ef6-",
    "Security and Confidentiality Software": "\u5e94\u7528\u8f6f\u4ef6-",
    "Embedded Application Software": "\u5e94\u7528\u8f6f\u4ef6-",
    "Education Software": "\u5e94\u7528\u8f6f\u4ef6-",
    "Game Software": "\u5e94\u7528\u8f6f\u4ef6-",
}

PRODUCT_TYPE_CN = {
    "Operating System": "\u64cd\u4f5c\u7cfb\u7edf",
    "Chinese Processing System": "\u4e2d\u6587\u5904\u7406\u7cfb\u7edf",
    "Network System": "\u7f51\u7edc\u7cfb\u7edf",
    "Embedded Operating System": "\u5d4c\u5165\u5f0f\u64cd\u4f5c\u7cfb\u7edf",
    "Other(System)": "\u5176\u4ed6",
    "Programming Language": "\u7a0b\u5e8f\u8bbe\u8ba1\u8bed\u8a00",
    "Database System Design": "\u6570\u636e\u5e93\u7cfb\u7edf\u8bbe\u8ba1",
    "Tools": "\u5de5\u5177\u8f6f\u4ef6",
    "Network Communication Software": "\u7f51\u7edc\u901a\u4fe1\u8f6f\u4ef6",
    "Middleware": "\u4e2d\u95f4\u4ef6",
    "Other(Support)": "\u5176\u4ed6",
    "Industry Management Software": "\u884c\u4e1a\u7ba1\u7406\u8f6f\u4ef6",
    "Office Software": "\u529e\u516c\u8f6f\u4ef6",
    "Pattern Recognition Software": "\u6a21\u5f0f\u8bc6\u522b\u8f6f\u4ef6",
    "Graphics Software": "\u56fe\u5f62\u56fe\u8c61\u8f6f\u4ef6",
    "Control Software": "\u63a7\u5236\u8f6f\u4ef6",
    "Network Application Software": "\u7f51\u7edc\u5e94\u7528\u8f6f\u4ef6",
    "Information Management Software": "\u4fe1\u606f\u7ba1\u7406\u8f6f\u4ef6",
    "Database Management Application Software": "\u6570\u636e\u5e93\u7ba1\u7406\u5e94\u7528\u8f6f\u4ef6",
    "Security and Confidentiality Software": "\u5b89\u5168\u4e0e\u4fdd\u5bc6\u8f6f\u4ef6",
    "Embedded Application Software": "\u5d4c\u5165\u5f0f\u5e94\u7528\u8f6f\u4ef6",
    "Education Software": "\u6559\u80b2\u8f6f\u4ef6",
    "Game Software": "\u6e38\u620f\u8f6f\u4ef6",
}


CATEGORY_ID_OPTIONS = {
    "01 \u64cd\u4f5c\u7cfb\u7edf",
    "02 \u5de5\u5177\u8f6f\u4ef6\u4e0e\u5e73\u53f0\u7cfb\u7edf",
    "03 \u4e2d\u95f4\u4ef6",
    "04 \u4fe1\u606f\u5b89\u5168",
    "05 \u5176\u5b83\u57fa\u7840\u8f6f\u4ef6",
    "06 \u4fe1\u606f\u901a\u8baf(ICT)",
    "07 \u6570\u5b57\u88c5\u5907",
    "08 \u533b\u7597\u8bbe\u5907",
    "09 \u5546\u7528\u53ca\u529e\u516c\u8bbe\u5907",
    "10 \u6570\u5b57\u7535\u89c6",
    "11 \u6c7d\u8f66\u7535\u5b50",
    "12 \u8ba1\u7b97\u673a\u53ca\u8bbe\u5907",
    "13 \u6d88\u8d39\u7535\u5b50",
    "14 \u4fe1\u606f\u5bb6\u7535",
    "15 \u5176\u4ed6\u901a\u8baf\u548c\u5de5\u4e1a",
    "16 \u529e\u516c\u548c\u7ba1\u7406",
    "17 \u4f01\u4e1a\u7ba1\u7406",
    "18 \u7535\u5b50\u653f\u52a1",
    "19 \u533b\u7597\u536b\u751f",
    "20 \u6559\u80b2",
    "21 \u5730\u7406\u4fe1\u606f",
    "22 \u91d1\u878d",
    "23 \u4ea4\u901a\u7269\u6d41",
    "24 \u6587\u5316\u521b\u610f",
    "25 \u5546\u8d38\u65c5\u6e38",
    "26 \u901a\u8baf\u7f51\u7edc\u670d\u52a1",
    "27 \u80fd\u6e90\u548c\u73af\u4fdd",
    "28 \u5efa\u7b51\u7269\u4e1a",
    "29 \u4e92\u8054\u7f51\u670d\u52a1",
    "30 \u5176\u4ed6\u8ba1\u7b97\u673a\u5e94\u7528\u8f6f\u4ef6\u548c\u4fe1\u606f\u670d\u52a1",
    "31 IC\u8bbe\u8ba1",
}

FIELD_PROMPTS = {
    "product__service_object": "从本段提炼服务对象，写成一句中文短语，仅说明面向人群/组织，不要写目的，80-120字。",
    "product__main_functions": "基于模块列表总结主要功能，80-120字，面向使用场景。",
    "product__tech_specs": "提炼3-5条技术指标式描述，中文短句，偏可验证特性。",
    "product__app_domain": "提取应用领域，中文短语，不带英文或括号。",
    "env__dev_platform": "概括开发平台，30-60字，避免硬件细节，只保留OS/主要工具/框架。",
    "env__run_platform": "概括运行平台，30-60字，仅说明运行操作系统/运行环境类型，不要列配置。",
    "env__hw_dev_platform": "提取开发硬件环境描述，保持原始配置要点。",
    "env__sw_dev_platform": "提取开发软件环境/工具，按“OS/工具/框架”顺序简述。",
    "env__memory_req": "提取内存要求，输出格式如 512MB。",
    "env__hardware_model": "提取适用机型，输出简短机型描述。",
    "env__language": "提取编程语言，使用原文名称，必须输出完整名字，不允许缩写。",
    "env__database": "提取数据库类型/版本，若多项取主库。",
    "env__os_version": "提取开发OS及版本，保持原文格式。",
    "env__server_soft": "输出服务器端应用软件列表，可多项，每项必须为软件名+版本号。",
    "env__client_soft": "输出客户端应用软件列表，可多项，每项必须为软件名+版本号。",
}

FIELD_TITLE_KEYWORDS = {
    "product__service_object": ["开发目的", "目标", "定位"],
    "product__main_functions": ["主要功能", "功能架构", "功能详述", "模块"],
    "product__func_list": ["主要功能", "功能详述", "模块"],
    "product__tech_specs": ["技术特点", "技术特性", "非功能性", "性能", "可靠性", "扩展性"],
    "product__app_domain": ["应用领域", "应用场景", "功能"],
    "env__dev_platform": ["开发环境", "开发平台"],
    "env__run_platform": ["运行环境", "部署环境", "软件运行环境"],
    "env__hw_dev_platform": ["开发环境", "硬件环境"],
    "env__sw_dev_platform": ["开发环境", "开发工具", "软件开发环境"],
    "env__memory_req": ["运行环境", "硬件", "硬件环境"],
    "env__hardware_model": ["运行环境", "硬件", "硬件环境"],
    "env__language": ["开发环境", "开发语言", "技术栈"],
    "env__database": ["数据库", "开发环境", "运行环境", "核心数据库"],
    "env__os_version": ["开发环境", "操作系统"],
    "env__server_soft": ["运行环境", "支撑软件", "软件运行环境"],
    "env__client_soft": ["运行环境", "客户端", "浏览器"],
}

FIELD_QUERIES = {
    "product__service_object": "服务对象 面向用户 目标群体 适用对象",
    "product__main_functions": "主要功能 功能架构 模块 功能描述",
    "product__func_list": "功能模块 主要功能 详述 模块名称",
    "product__tech_specs": "技术指标 技术特点 高性能 高可靠 可扩展 安全",
    "product__app_domain": "应用领域 应用场景 行业",
    "env__dev_platform": "开发平台 开发环境 工具 框架",
    "env__run_platform": "运行平台 运行环境 部署平台",
    "env__hw_dev_platform": "开发硬件 环境 配置",
    "env__sw_dev_platform": "开发工具 框架 IDE OS",
    "env__memory_req": "内存要求 内存",
    "env__hardware_model": "适用机型 硬件 设备",
    "env__language": "开发语言 编程语言 前端 后端",
    "env__database": "数据库 类型 版本 MySQL PostgreSQL Oracle",
    "env__os_version": "操作系统 版本",
    "env__server_soft": "服务器 软件 中间件 数据库 运行支撑",
    "env__client_soft": "客户端 软件 浏览器",
}

SERVER_MODEL_POOL = [
    {
        "model": "Dell PowerEdge R750",
        "config": "CPU：Intel Xeon Silver 4314\n内存：64GB\n硬盘：2TB SSD",
    },
    {
        "model": "HPE ProLiant DL380 Gen10",
        "config": "CPU：Intel Xeon Gold 5218\n内存：128GB\n硬盘：4TB SSD",
    },
    {
        "model": "Lenovo ThinkSystem SR650",
        "config": "CPU：Intel Xeon Silver 4216\n内存：64GB\n硬盘：2TB SSD",
    },
    {
        "model": "Inspur NF5280M5",
        "config": "CPU：Intel Xeon Silver 4210\n内存：64GB\n硬盘：2TB SSD",
    },
    {
        "model": "Huawei FusionServer 2288H V5",
        "config": "CPU：Intel Xeon Gold 6230\n内存：128GB\n硬盘：4TB SSD",
    },
    {
        "model": "Dell PowerEdge R740",
        "config": "CPU：Intel Xeon Gold 5220\n内存：128GB\n硬盘：4TB SSD",
    },
    {
        "model": "HPE ProLiant DL360 Gen10",
        "config": "CPU：Intel Xeon Silver 4210\n内存：64GB\n硬盘：2TB SSD",
    },
    {
        "model": "Lenovo ThinkSystem SR630",
        "config": "CPU：Intel Xeon Silver 4214\n内存：64GB\n硬盘：2TB SSD",
    },
]

CLIENT_MODEL_POOL = [
    {
        "model": "Lenovo ThinkPad T14 Gen 4",
        "config": "CPU：Intel Core i7-1360P\n内存：16GB\n硬盘：1TB SSD",
    },
    {
        "model": "Dell Precision 3660",
        "config": "CPU：Intel Core i7-13700\n内存：32GB\n硬盘：1TB SSD",
    },
    {
        "model": "HP EliteDesk 800 G9",
        "config": "CPU：Intel Core i5-13500\n内存：16GB\n硬盘：512GB SSD",
    },
    {
        "model": "Lenovo ThinkCentre M90t",
        "config": "CPU：Intel Core i7-12700\n内存：16GB\n硬盘：1TB SSD",
    },
    {
        "model": "Dell OptiPlex 7010",
        "config": "CPU：Intel Core i5-13500\n内存：16GB\n硬盘：512GB SSD",
    },
    {
        "model": "ASUS ProArt Studiobook 16",
        "config": "CPU：Intel Core i9-13980HX\n内存：32GB\n硬盘：1TB SSD",
    },
    {
        "model": "Huawei MateBook D16",
        "config": "CPU：Intel Core i5-12450H\n内存：16GB\n硬盘：512GB SSD",
    },
    {
        "model": "Apple MacBook Pro 14",
        "config": "CPU：Apple M2 Pro\n内存：16GB\n硬盘：512GB SSD",
    },
]

SERVER_OS_POOL = [
    "Windows Server 2019",
    "Windows Server 2022",
    "Ubuntu Server 20.04 LTS",
    "Ubuntu Server 22.04 LTS",
    "Red Hat Enterprise Linux 8.8",
    "Red Hat Enterprise Linux 9.2",
    "CentOS Stream 9",
    "SUSE Linux Enterprise Server 15 SP5",
    "Debian 11",
    "Debian 12",
    "银河麒麟高级服务器操作系统 V10",
    "统信UOS服务器版 V20",
    "中标麒麟高级服务器操作系统 V7",
]

CLIENT_OS_POOL = [
    "Windows 10 专业版 22H2",
    "Windows 11 专业版 23H2",
    "macOS Ventura 13",
    "macOS Sonoma 14",
    "Ubuntu 22.04 LTS",
    "Ubuntu 24.04 LTS",
]


def ensure_local_run() -> bool:
    repo_root = Path(__file__).resolve().parents[1]
    if not (repo_root / "pyproject.toml").exists():
        print("Local-only mode: running from a global install is disabled.")
        return False
    cwd = Path.cwd().resolve()
    if not cwd.is_relative_to(repo_root):
        print(f"Run from the repository folder: {repo_root}")
        return False
    return True


def is_abs_path(path_value: str) -> bool:
    if Path(path_value).is_absolute():
        return True
    return bool(re.match(r"^[A-Za-z]:[\\/]", path_value))


def load_template_paths() -> dict[str, Path] | None:
    missing = []
    template_paths: dict[str, Path] = {}
    for key, env_key in TEMPLATE_ENV_MAP.items():
        value = os.getenv(env_key, "").strip()
        if not value:
            missing.append(env_key)
            continue
        if not is_abs_path(value):
            print(f"Template path must be absolute: {env_key}={value}")
            return None
        path = Path(value).expanduser()
        template_paths[key] = path
    if missing:
        print(f"Missing template path(s) in env: {', '.join(missing)}")
        return None
    return template_paths


def read_file_content(file_path: Path, max_chars: int = 40000) -> str:
    ext = file_path.suffix.lower()
    text = ""
    try:
        if ext == ".md":
            text = file_path.read_text(encoding="utf-8")[:max_chars]
    except Exception as exc:
        logging.error("Failed to read %s: %s", file_path, exc)
        return ""
    return text


def init_client(api_key: str, base_url: str | None) -> OpenAI:
    if not api_key:
        raise ValueError("Missing API key. Use --api-key or set OPENAI_API_KEY.")
    if base_url:
        return OpenAI(api_key=api_key, base_url=base_url)
    return OpenAI(api_key=api_key)


def call_llm_json(
    client: OpenAI,
    model: str,
    system_prompt: str,
    user_prompt: str,
    temperature: float = 0.1,
) -> dict:
    try:
        response = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            response_format={"type": "json_object"},
            temperature=temperature,
        )
        return json.loads(response.choices[0].message.content)
    except Exception as exc:
        logging.error("LLM request failed: %s", exc)
        return {}


def _chinese_numeral_to_int(value: str) -> int | None:
    mapping = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "七": 7, "八": 8, "九": 9, "十": 10}
    if value in mapping:
        return mapping[value]
    if len(value) == 2 and value.startswith("十"):
        return 10 + mapping.get(value[1], 0)
    if len(value) == 2 and value.endswith("十"):
        return mapping.get(value[0], 0) * 10
    if len(value) == 3 and value[1] == "十":
        return mapping.get(value[0], 0) * 10 + mapping.get(value[2], 0)
    return None


def build_section_map(text: str) -> dict[str, dict[str, str]]:
    sections: dict[str, dict[str, str]] = {}
    current_id = None
    current_title = None
    for line in text.splitlines():
        line_stripped = line.strip()
        if not line_stripped:
            if current_id is not None:
                sections[current_id]["content"] += line + "\n"
            continue
        heading_match = re.match(r"^#+\s*(.+)$", line_stripped)
        if heading_match:
            heading = heading_match.group(1).strip()
            num_match = re.match(r"^(\d+(?:\.\d+)+)\s+(.+)$", heading)
            if num_match:
                current_id = num_match.group(1)
                current_title = num_match.group(2).strip()
                sections.setdefault(current_id, {"title": current_title, "content": ""})
                continue
            chapter_match = re.match(r"^第([一二三四五六七八九十]+)章\s+(.+)$", heading)
            if chapter_match:
                num = _chinese_numeral_to_int(chapter_match.group(1))
                current_id = str(num) if num else chapter_match.group(1)
                current_title = chapter_match.group(2).strip()
                sections.setdefault(current_id, {"title": current_title, "content": ""})
                continue
            current_id = f"title:{heading}"
            current_title = heading
            sections.setdefault(current_id, {"title": current_title, "content": ""})
            continue
        if current_id is not None:
            sections[current_id]["content"] += line + "\n"
    return {
        key: {"title": value.get("title", ""), "content": value.get("content", "").strip()}
        for key, value in sections.items()
        if value.get("content", "").strip()
    }


@dataclass
class Chunk:
    text: str
    section_id: str | None = None
    section_title: str | None = None


def split_into_chunks(full_text: str, max_chars: int = 600, overlap: int = 80) -> list[Chunk]:
    paragraphs = [p.strip() for p in re.split(r"\n\s*\n", full_text) if p.strip()]
    chunks: list[Chunk] = []
    buf = ""
    for paragraph in paragraphs:
        if len(buf) + len(paragraph) + 2 <= max_chars:
            buf += ("\n\n" + paragraph) if buf else paragraph
        else:
            if buf:
                chunks.append(Chunk(text=buf))
            buf = (buf[-overlap:] + "\n\n" + paragraph) if overlap > 0 else paragraph
    if buf:
        chunks.append(Chunk(text=buf))
    return chunks


def build_section_chunks(section_map: dict[str, dict[str, str]], max_chars: int = 900) -> list[Chunk]:
    chunks: list[Chunk] = []
    for section_id, section in section_map.items():
        title = section.get("title", "")
        content = section.get("content", "").strip()
        if not content:
            continue
        if len(content) <= max_chars:
            chunks.append(Chunk(text=content, section_id=section_id, section_title=title))
        else:
            for sub in split_into_chunks(content, max_chars=max_chars, overlap=120):
                sub.section_id = section_id
                sub.section_title = title
                chunks.append(sub)
    return chunks


MODULE_TITLE_WORDS = ["模块", "子系统", "平台", "中心", "后台", "工作台", "服务", "系统", "应用", "调度", "管理"]
MODULE_TEXT_HINTS = ["功能模块", "模块包括", "系统包括", "系统由", "主要模块", "功能组成"]
FALLBACK_MIN_MODULES = 3


def is_module_candidate(chunk_title: str | None, chunk_text: str) -> bool:
    title = (chunk_title or "").strip()
    content = (chunk_text or "").strip()
    if any(word in title for word in MODULE_TITLE_WORDS):
        return True
    if any(hint in content for hint in MODULE_TEXT_HINTS):
        return True
    return False


def clean_module_title(title: str) -> str:
    cleaned = title.strip()
    cleaned = re.sub(r"^\d+(\.\d+)*\s*[\.\、]?\s*", "", cleaned)
    cleaned = re.sub(r"（.*?）|\(.*?\)", "", cleaned).strip()
    cleaned = cleaned.replace("“", "").replace("”", "").replace('"', "")
    cleaned = re.sub(r"[：:，,。\.]+$", "", cleaned).strip()
    return cleaned


def collect_module_candidates_from_chunks(chunks: list[Chunk]) -> list[str]:
    titles: list[str] = []
    for chunk in chunks:
        if is_module_candidate(chunk.section_title, chunk.text):
            if chunk.section_title:
                titles.append(clean_module_title(chunk.section_title))
    seen: set[str] = set()
    result: list[str] = []
    for title in titles:
        if len(title) < 4:
            continue
        if title not in seen:
            seen.add(title)
            result.append(title)
    return result


def build_module_query(title: str) -> str:
    return f"{title} 功能 支持 用于 提供 实现 可以"


def normalize_func_desc(desc: str, max_len: int = 20) -> str:
    cleaned = re.sub(r"[，。；、,.;:：]", "", desc)
    cleaned = re.sub(r"\s+", "", cleaned)
    if not cleaned:
        return ""
    if not cleaned.startswith("可以"):
        cleaned = "可以" + cleaned
    if len(cleaned) > max_len:
        cleaned = cleaned[:max_len]
    return cleaned


def normalize_llm_func_items_from_response(response: dict) -> list[dict[str, str]]:
    items = response.get("items") if isinstance(response, dict) else None
    if not isinstance(items, list):
        return []
    return normalize_llm_func_items(items)


def extract_func_items(
    full_text: str,
    section_chunks: list[Chunk],
    client: OpenAI,
    model: str,
) -> list[dict[str, str]]:
    if not section_chunks:
        return []
    retriever = BM25Retriever(section_chunks)
    evidences = build_module_evidences(section_chunks, retriever)
    if len(evidences) < FALLBACK_MIN_MODULES:
        return []
    system_prompt = (
        "你是软件测评文档编写助手。请根据每个模块的证据文本，输出“产品测试功能表”所需内容。"
        "输出为 JSON 对象，包含 items 数组。"
        "每个元素包含字段：一级功能、功能描述。"
        "一级功能使用给定 module_title（可清理但不要新增模块）。"
        "功能描述必须以“可以”开头，仅一句话，25~60个中文字符。"
        "不要出现编号、引号、冒号，不要出现“本模块/该模块/它”等代词。"
        "只输出 JSON，不要解释。"
    )
    user_prompt = json.dumps(evidences, ensure_ascii=False)
    response = call_llm_json(client, model, system_prompt, user_prompt, temperature=0.2)
    return normalize_llm_func_items_from_response(response)


def filter_chunks_by_title_keywords(field: str, chunks: list[Chunk], min_keep: int = 6) -> list[Chunk]:
    keywords = FIELD_TITLE_KEYWORDS.get(field)
    if not keywords:
        return chunks
    filtered = []
    for chunk in chunks:
        title = chunk.section_title or ""
        if any(keyword in title for keyword in keywords):
            filtered.append(chunk)
    return filtered if len(filtered) >= min_keep else chunks


def tokenize(text: str) -> list[str]:
    text = re.sub(r"\s+", " ", text)
    tokens = re.findall(r"[\u4e00-\u9fff]|[A-Za-z0-9]+", text)
    return [token.lower() for token in tokens if token.strip()]


class BM25Retriever:
    def __init__(self, chunks: list[Chunk]):
        self.chunks = chunks
        self.corpus_tokens = [tokenize(chunk.text) for chunk in chunks]
        try:
            from rank_bm25 import BM25Okapi

            self.bm25 = BM25Okapi(self.corpus_tokens)
            self.use_bm25 = True
        except Exception:
            self.bm25 = None
            self.use_bm25 = False

    def retrieve(self, query: str, top_k: int = 4) -> list[tuple[Chunk, float]]:
        if not self.chunks:
            return []
        query_tokens = tokenize(query)
        if self.use_bm25 and self.bm25 is not None:
            scores = self.bm25.get_scores(query_tokens)
        else:
            scores = []
            for tokens in self.corpus_tokens:
                score = sum(tokens.count(token) for token in query_tokens)
                scores.append(float(score))
        ranked = sorted(enumerate(scores), key=lambda x: x[1], reverse=True)[:top_k]
        return [(self.chunks[i], float(score)) for i, score in ranked]


def retrieve_evidence_for_module(title: str, retriever: BM25Retriever, top_k: int = 2) -> str:
    results = retriever.retrieve(build_module_query(title), top_k=top_k)
    parts = []
    for chunk, _score in results:
        header = chunk.section_title or ""
        parts.append(f"{header}\n{chunk.text}".strip())
    return "\n\n---\n\n".join(parts)


def build_module_evidences(chunks: list[Chunk], retriever: BM25Retriever) -> list[dict[str, str]]:
    titles = collect_module_candidates_from_chunks(chunks)
    evidences: list[dict[str, str]] = []
    for title in titles:
        evidence = retrieve_evidence_for_module(title, retriever, top_k=2)
        if len(evidence) < 80:
            continue
        evidences.append({"module_title": title, "evidence": evidence})
    return evidences


def normalize_llm_func_items(items: list[dict[str, str]]) -> list[dict[str, str]]:
    normalized: list[dict[str, str]] = []
    for item in items:
        title = clean_module_title(str(item.get("一级功能") or "").strip())
        desc = str(item.get("功能描述") or "").strip()
        if not title:
            continue
        desc = re.sub(r"\s+", "", desc)
        desc = re.sub(r"^[\d\.\、]+", "", desc)
        desc = desc.replace("：", "，").replace(":", "，")
        desc = desc.replace("“", "").replace("”", "").replace('"', "")
        if not desc.startswith("可以"):
            desc = "可以" + desc
        if len(desc) > 60:
            desc = desc[:60] + "…"
        desc = re.split(r"[。]", desc)[0]
        normalized.append({"name": title, "desc": desc.rstrip("，,、")})
    seen: set[str] = set()
    final_items: list[dict[str, str]] = []
    for item in normalized:
        if item["name"] in seen:
            continue
        seen.add(item["name"])
        final_items.append(item)
    return final_items


def get_context_for_field(
    field: str,
    full_text: str,
    section_chunks: list[Chunk],
    full_chunks: list[Chunk],
    top_k: int = 4,
    min_chars: int = 800,
) -> str:
    base_chunks = section_chunks if section_chunks else full_chunks
    candidate_chunks = filter_chunks_by_title_keywords(field, base_chunks)
    if not candidate_chunks:
        candidate_chunks = full_chunks or base_chunks
    query = FIELD_QUERIES.get(field, field.replace("__", " "))
    retriever = BM25Retriever(candidate_chunks)
    results = retriever.retrieve(query, top_k=top_k)
    parts = []
    for chunk, _score in results:
        header = ""
        if chunk.section_id or chunk.section_title:
            header = f"[{chunk.section_id or ''} {chunk.section_title or ''}]".strip()
        parts.append(f"{header}\n{chunk.text}".strip())
    context = "\n\n---\n\n".join(parts).strip()
    if len(context) < min_chars:
        context = (context + "\n\n---\n\n" + full_text[:4000]).strip()
    return context


def extract_fields_by_prompt(
    client: OpenAI,
    model: str,
    full_text: str,
    section_chunks: list[Chunk],
    full_chunks: list[Chunk],
    field_prompts: dict[str, str],
) -> dict:
    results: dict[str, object] = {}
    for field, prompt in field_prompts.items():
        system_prompt = (
            "你只需输出JSON对象，且只包含一个字段。字段名必须严格等于要求的字段名，"
            f"字段名为 {field}。{prompt}"
        )
        context = get_context_for_field(field, full_text, section_chunks, full_chunks)
        user_prompt = f"章节内容如下：\n{context}"
        data = call_llm_json(client, model, system_prompt, user_prompt, temperature=0.2)
        if isinstance(data, dict) and field in data:
            results[field] = data[field]
    return results


def sanitize_html_breaks(value: object) -> object:
    if isinstance(value, str):
        return (
            value.replace("<br />", "\n")
            .replace("<br/>", "\n")
            .replace("<br>", "\n")
        )
    if isinstance(value, list):
        return [sanitize_html_breaks(item) for item in value]
    if isinstance(value, dict):
        return {k: sanitize_html_breaks(v) for k, v in value.items()}
    return value


def sanitize_data(data: dict) -> dict:
    data = sanitize_html_breaks(data)
    data = normalize_list_fields(data)
    data = normalize_server_client_fields(data)
    return data


def _to_workday(value: "date") -> "date":
    while value.weekday() >= 5:
        value = value - timedelta(days=1)
    return value


def _subtract_months(value: "date", months: int) -> "date":
    year = value.year
    month = value.month - months
    while month <= 0:
        year -= 1
        month += 12
    last_day = calendar.monthrange(year, month)[1]
    day = min(value.day, last_day)
    return value.replace(year=year, month=month, day=day)


def normalize_list_fields(data: dict) -> dict:
    for key, value in list(data.items()):
        if key == "product__func_list":
            continue
        if isinstance(value, list):
            data[key] = "\n".join(str(item) for item in value if item is not None)
    return data


def normalize_server_client_fields(data: dict) -> dict:
    def split_pair(value: str) -> tuple[str, str] | None:
        if not value:
            return None
        pattern = re.compile(r"(服务器端|服务器)[:：]\s*(.+?)\s*(客户端)[:：]\s*(.+)", re.DOTALL)
        match = pattern.search(value)
        if not match:
            return None
        server_part = match.group(2).strip()
        client_part = match.group(4).strip()
        return server_part, client_part

    pairs = [
        ("env__server_os", "env__client_os"),
        ("env__server_soft", "env__client_soft"),
        ("env__server_config", "env__client_config"),
        ("env__server_model", "env__client_model"),
    ]
    for server_key, client_key in pairs:
        combined = data.get(server_key, "")
        split = split_pair(str(combined))
        if split:
            data[server_key], data[client_key] = split
            continue
        combined = data.get(client_key, "")
        split = split_pair(str(combined))
        if split:
            data[server_key], data[client_key] = split
    return data


def pick_random_models() -> tuple[dict, dict]:
    server_choice = random.choice(SERVER_MODEL_POOL)
    client_choice = random.choice(CLIENT_MODEL_POOL)
    return server_choice, client_choice


def pick_random_os() -> tuple[str, str]:
    server_os = random.choice(SERVER_OS_POOL)
    client_os = random.choice(CLIENT_OS_POOL)
    return server_os, client_os


def parse_markdown_tables(text: str) -> list[tuple[list[str], list[list[str]]]]:
    tables: list[tuple[list[str], list[list[str]]]] = []
    lines = text.splitlines()
    i = 0
    while i < len(lines) - 1:
        header = lines[i]
        sep = lines[i + 1]
        if "|" in header and "|" in sep and set(sep.strip()) <= {"|", "-", ":", " "}:
            header_cells = [cell.strip() for cell in header.strip("|").split("|")]
            rows: list[list[str]] = []
            i += 2
            while i < len(lines) and "|" in lines[i]:
                row_cells = [cell.strip() for cell in lines[i].strip("|").split("|")]
                if any(row_cells):
                    rows.append(row_cells)
                i += 1
            tables.append((header_cells, rows))
            continue
        i += 1
    return tables


def extract_app_from_text(text: str) -> dict:
    data: dict[str, str] = {}
    name_match = re.search(r"软件全称[:：]\s*(.+)", text)
    version_match = re.search(r"版\s*本\s*号[:：]\s*(.+)", text)
    if name_match:
        data["app__name"] = name_match.group(1).strip()
    if version_match:
        data["app__version"] = version_match.group(1).strip()
    return data


def extract_func_list_from_text(text: str) -> dict:
    pattern = re.compile(r"^#{1,6}\s*3\.3\.\d+\s*([^\n]+)", re.MULTILINE)
    modules = [name.strip() for name in pattern.findall(text) if name.strip()]
    if not modules:
        return {}
    return {
        "product__func_list": [
            {"name": module, "desc": ""} for module in modules
        ]
    }


def extract_env_from_tables(text: str) -> dict:
    tables = parse_markdown_tables(text)
    data: dict[str, str] = {}
    for header, rows in tables:
        if "类别" not in header:
            continue
        config_idx = None
        for idx, name in enumerate(header):
            if "配置" in name:
                config_idx = idx
                break
        if config_idx is None:
            continue
        category_idx = header.index("类别")
        for row in rows:
            if category_idx >= len(row) or config_idx >= len(row):
                continue
            category = row[category_idx]
            value = row[config_idx]
            if not category or not value:
                continue
            if "开发硬件环境" in category:
                data["env__hw_dev_platform"] = value
                data["tech__hardware_dev"] = value
            elif "开发该软件的操作系统" in category or "开发操作系统" in category:
                data["tech__os_dev"] = value
            elif "编程语言" in category:
                data["env__language"] = value
                data["tech__language"] = value
            elif "开发环境" in category or "开发工具" in category:
                data["env__sw_dev_platform"] = value
                data["tech__dev_tools"] = value
            elif "运行硬件环境" in category:
                data["tech__hardware_run"] = value
                data["env__server_config"] = value
                data["env__client_config"] = value
            elif "运行平台" in category or "运行操作系统" in category:
                data["env__server_os"] = value
                data["env__client_os"] = value
                data["tech__os_run"] = value
            elif "运行支撑" in category or "支持软件" in category:
                data["env__server_soft"] = value
                data["env__client_soft"] = value
                data["tech__run_support"] = value
    return data


def extract_rule_data(text: str) -> dict:
    data: dict[str, object] = {}
    data.update(extract_app_from_text(text))
    data.update(extract_func_list_from_text(text))
    data.update(extract_env_from_tables(text))
    return data


def load_yaml_config(path: Path) -> dict:
    if not path.exists():
        logging.warning("Config file not found: %s", path)
        return {}
    try:
        with path.open("r", encoding="utf-8") as f:
            return yaml.safe_load(f) or {}
    except Exception as exc:
        logging.error("Failed to load config: %s", exc)
        return {}


def select_preset(config: dict, choice: str | None) -> dict:
    presets = config.get("presets", [])
    if not presets:
        return {}
    if choice:
        for item in presets:
            if item.get("label") == choice:
                return item
    default_label = config.get("preset_choice")
    if default_label:
        for item in presets:
            if item.get("label") == default_label:
                return item
    return presets[0]


def resolve_contact_info(config: dict, preset: dict) -> dict:
    if preset and preset.get("contact_info"):
        return preset["contact_info"]
    return config.get("contact_info", {})


def sanitize_filename(value: str) -> str:
    unsafe = '<>:"/\\|?*'
    cleaned = "".join(ch for ch in value if ch not in unsafe).strip()
    return cleaned.replace(" ", "")


def build_filename(prefix: str, software_name: str, version: str, suffix: str) -> str:
    name = sanitize_filename(software_name) or "\u672a\u547d\u540d\u8f6f\u4ef6"
    ver = sanitize_filename(version) or "\u672a\u6807\u6ce8\u7248\u672c"
    return f"{prefix}-{name}-{ver}{suffix}"


def build_copyright_filename(software_name: str, version: str) -> str:
    name = sanitize_filename(software_name) or "\u672a\u547d\u540d\u8f6f\u4ef6"
    ver = sanitize_filename(version) or "\u672a\u6807\u6ce8\u7248\u672c"
    return f"{name}-{ver}-\u8f6f\u8457\u7533\u8bf7\u8868.docx"


def ensure_short_name(name: str) -> str:
    if not name:
        return "无"
    clean = re.sub(r"\s+", "", name)
    if len(clean) <= 6:
        return clean
    return clean[:6]


def normalize_product_type(value: str) -> str:
    if not value:
        return "应用软件-信息管理软件"
    if "Other" in value or "其他" in value:
        return "应用软件-信息管理软件"
    for key, cn_value in PRODUCT_TYPE_CN.items():
        if cn_value in value or key in value:
            prefix = PRODUCT_TYPE_PREFIX.get(key, "应用软件-")
            return f"{prefix}{cn_value}"
    return value


def normalize_category_id(value: str) -> str:
    value = str(value).strip()
    if value in CATEGORY_ID_OPTIONS:
        return value
    return "30 其它计算机应用软件和信息服务"


def normalize_assessment_data(data: dict) -> dict:
    if not data:
        return {}
    for key in CELL_MAP_TEXT:
        value = data.get(key)
        if isinstance(value, list):
            data[key] = "\n".join(str(item) for item in value if item is not None)
    defaults = {
        "assess__support_floppy": False,
        "assess__support_sound": False,
        "assess__support_cdrom": False,
        "assess__support_gpu": False,
        "assess__support_other": False,
        "assess__is_self_dev": True,
        "assess__has_docs": True,
        "assess__has_source": True,
    }
    for key, value in defaults.items():
        if key not in data:
            data[key] = value
        elif isinstance(data[key], str):
            data[key] = data[key].strip().lower() in {"true", "yes", "1"}

    workload = str(data.get("assess__workload", "")).strip()
    if workload:
        numbers = re.findall(r"\d+", workload)
        people = int(numbers[0]) if numbers else 3
        months = int(numbers[1]) if len(numbers) > 1 else 3
        people = max(1, min(people, 6))
        months = max(1, months)
        data["assess__workload"] = f"{people}\u4eba*{months}\u6708"
    else:
        data["assess__workload"] = "3\u4eba*3\u6708"

    memory = str(data.get("env__memory_req", "")).strip()
    if memory:
        digits = "".join(ch for ch in memory if ch.isdigit())
        if digits:
            data["env__memory_req"] = f"{digits}MB"

    if data.get("env__soft_scale") not in {"\u5927", "\u4e2d", "\u5c0f"}:
        data["env__soft_scale"] = "\u4e2d"

    data["app__category_assess"] = normalize_category_id(data.get("app__category_assess", ""))
    data["app__product_type_text"] = normalize_product_type(data.get("app__product_type_text", ""))

    if not str(data.get("env__database", "")).strip():
        data["env__database"] = "PostgreSQL 13"
    if not str(data.get("env__os_version", "")).strip():
        data["env__os_version"] = "Ubuntu 22.04 LTS"

    return data


def merge_missing(base: dict, updates: dict) -> dict:
    for key, value in updates.items():
        current = base.get(key)
        if key not in base or current in (None, "") or current == []:
            base[key] = value
    return base


def derive_fields(data: dict) -> dict:
    if not data.get("tech__source_lines"):
        data["tech__source_lines"] = "15000"
    if not data.get("app__short_name"):
        data["app__short_name"] = "无"

    func_list = data.get("product__func_list", [])
    if func_list:
        for item in func_list:
            if not item.get("desc"):
                name = item.get("name", "")
                if name:
                    short = name[:6]
                    item["desc"] = f"可以{short}管理"

    if not data.get("product__main_functions") and func_list:
        names = [item.get("name") for item in func_list if item.get("name")]
        if names:
            data["product__main_functions"] = "、".join(names[:5]) + "等功能"

    main_functions = data.get("product__main_functions", "")
    if main_functions:
        data["tech__main_functions"] = f"系统主要功能包括：{main_functions}"

    tech_specs = data.get("product__tech_specs", "")
    if tech_specs:
        if isinstance(tech_specs, str) and "；" in tech_specs:
            specs_text = tech_specs.replace("；", "，")
        else:
            specs_text = tech_specs
        data["tech__features"] = f"技术特点体现在：{specs_text}"

    service_object = data.get("product__service_object", "")
    if service_object:
        data["tech__dev_purpose"] = f"开发目的在于服务{service_object}，解决其业务需求。"

    language = data.get("env__language", "")
    if language:
        data["tech__language"] = language

    hw_dev = data.get("env__hw_dev_platform", "")
    if hw_dev:
        data["tech__hardware_dev"] = f"开发硬件环境：{hw_dev}"

    dev_tools = data.get("env__sw_dev_platform", "")
    if dev_tools:
        data["tech__dev_tools"] = f"开发工具与平台：{dev_tools}"

    os_dev = data.get("env__os_version") or data.get("env__os", "")
    if os_dev:
        data["tech__os_dev"] = os_dev

    server_os = data.get("env__server_os", "")
    client_os = data.get("env__client_os", "")
    if server_os or client_os:
        if server_os and client_os:
            data["tech__os_run"] = f"服务器端：{server_os}；客户端：{client_os}"
        else:
            data["tech__os_run"] = server_os or client_os

    server_soft = data.get("env__server_soft", "")
    client_soft = data.get("env__client_soft", "")
    if server_soft or client_soft:
        if server_soft and client_soft:
            data["tech__run_support"] = f"服务器端：{server_soft}；客户端：{client_soft}"
        else:
            data["tech__run_support"] = server_soft or client_soft

    server_choice, client_choice = pick_random_models()
    data["env__server_model"] = server_choice["model"]
    data["env__server_config"] = server_choice["config"]
    data["env__client_model"] = client_choice["model"]
    data["env__client_config"] = client_choice["config"]

    server_os, client_os = pick_random_os()
    data["env__server_os"] = server_os
    data["env__client_os"] = client_os

    if data.get("env__language") and not data.get("env__dev_lang"):
        data["env__dev_lang"] = data.get("env__language")

    if not data.get("env__os"):
        data["env__os"] = "Windows, Linux, macOS"

    data["env__memory_req"] = "2048MB"

    if not data.get("env__dev_platform") and data.get("env__sw_dev_platform"):
        data["env__dev_platform"] = data.get("env__sw_dev_platform")

    if data.get("env__os") and not data.get("env__run_platform"):
        data["env__run_platform"] = data.get("env__os")

    if not data.get("copyright__completion_date"):
        data["copyright__completion_date"] = _to_workday(date.today() - timedelta(days=14)).strftime("%Y-%m-%d")

    try:
        base_date = datetime.strptime(data["copyright__completion_date"], "%Y-%m-%d").date()
    except Exception:
        base_date = None

    if base_date and not data.get("assess__completion_date"):
        data["assess__completion_date"] = base_date.strftime("%Y/%m/%d")

    if base_date and not data.get("assess__dev_date"):
        min_dev_date = _subtract_months(base_date, 8)
        dev_date = _to_workday(min_dev_date)
        if dev_date > min_dev_date:
            dev_date = _to_workday(min_dev_date)
        data["assess__dev_date"] = dev_date.strftime("%Y/%m/%d")

    return data


def fill_func_table(template_path: Path, output_path: Path, data: dict) -> bool:
    func_list = data.get("product__func_list")
    if not func_list:
        return False
    context = {
        "software_name": data.get("app__name", ""),
        "func_list": func_list,
    }
    tpl = DocxTemplate(template_path)
    tpl.render(context)
    tpl.save(output_path)
    return True


def build_holder_context(contact_info: dict) -> dict:
    return {
        "holder__name": contact_info.get("owner", ""),
        "holder__address": contact_info.get("address", ""),
        "holder__zip_code": contact_info.get("zip_code", ""),
        "holder__contact_name": contact_info.get("contact_name", ""),
        "holder__contact_mobile": contact_info.get("contact_mobile", ""),
        "holder__contact_email": contact_info.get("contact_email", ""),
        "holder__contact_landline": contact_info.get("contact_landline", ""),
        "holder__tech_contact_name": contact_info.get("tech_contact_name", ""),
        "holder__tech_contact_mobile": contact_info.get("tech_contact_mobile", ""),
    }


def fill_reg_table(template_path: Path, output_path: Path, data: dict, contact_info: dict) -> bool:
    if not data:
        return False
    holder_context = build_holder_context(contact_info)
    replacements = {
        "{{app__name}}": data.get("app__name", ""),
        "{{app__version}}": data.get("app__version", ""),
        "{{app__short_name}}": data.get("app__short_name", ""),
        "{{env__dev_platform}}": data.get("env__dev_platform") or data.get("env__sw_dev_platform", ""),
        "{{env__dev_lang}}": data.get("env__dev_lang") or data.get("env__language", ""),
        "{{env__run_platform}}": data.get("env__run_platform") or data.get("env__os", ""),
        "{{product__app_domain}}": data.get("product__app_domain", ""),
    }
    for key, value in holder_context.items():
        replacements[f"{{{{{key}}}}}"] = value

    doc = DocxWriter(template_path)
    docx_replace_text(doc, replacements)
    doc.save(output_path)
    return True


def set_checkbox(ws, cell_coord: str, checked: bool) -> None:
    target = CHECKED_SYMBOL if checked else UNCHECKED_SYMBOL
    val = ws[cell_coord].value
    if val and isinstance(val, str):
        ws[cell_coord] = val.replace("\u25a1", target).replace("\u2610", target).replace("\u2611", target)
    else:
        ws[cell_coord] = target


def set_text_cell(ws, coord: str, value: str, wrap: bool) -> None:
    cell = ws[coord]
    cell.value = value or ""
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")


def fill_assessment_excel(template_path: Path, output_path: Path, data: dict) -> bool:
    if not data:
        return False
    wb = load_workbook(template_path)
    sheets = wb.worksheets

    for key, (sheet_idx, coord) in CELL_MAP_TEXT.items():
        value = data.get(key, "")
        set_text_cell(sheets[sheet_idx], coord, value, wrap=key in WRAP_TEXT_KEYS)

    for key, (sheet_idx, coord) in CELL_MAP_CHECKBOX.items():
        set_checkbox(sheets[sheet_idx], coord, data.get(key) is True)

    mode = data.get("assess__product_mode_val", "pure")
    sheet_idx, coord = CELL_MODE_PURE if mode != "embedded" else CELL_MODE_EMBEDDED
    set_checkbox(sheets[sheet_idx], coord, True)

    wb.save(output_path)
    return True


def fill_env_table(template_path: Path, output_path: Path, data: dict) -> bool:
    if not data:
        return False
    replacements = {
        "{{env__server_os}}": data.get("env__server_os", ""),
        "{{env__server_soft}}": data.get("env__server_soft", ""),
        "{{env__server_model}}": data.get("env__server_model", ""),
        "{{env__server_config}}": data.get("env__server_config", ""),
        "{{env__server_id}}": data.get("env__server_id", "\u5382\u5546\u8bbe\u5907"),
        "{{env__client_os}}": data.get("env__client_os", ""),
        "{{env__client_soft}}": data.get("env__client_soft", ""),
        "{{env__client_model}}": data.get("env__client_model", ""),
        "{{env__client_config}}": data.get("env__client_config", ""),
        "{{env__client_id}}": data.get("env__client_id", "\u5382\u5546\u8bbe\u5907"),
    }
    doc = DocxWriter(template_path)
    docx_replace_text(doc, replacements)
    doc.save(output_path)
    return True


def ensure_output_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def collect_inputs(inputs: list[str]) -> list[Path]:
    files: list[Path] = []
    for entry in inputs:
        path = Path(entry)
        if path.is_dir():
            for file_path in path.iterdir():
                if file_path.suffix.lower() in {".md"}:
                    files.append(file_path)
        elif path.is_file() and path.suffix.lower() in {".md"}:
            files.append(path)
    return files


def _prompt_choice(prompt: str, options: Iterable[str], default: str | None = None) -> str:
    opts = list(options)
    if default:
        prompt = f"{prompt} [{default}]"
    while True:
        value = input(f"{prompt}: ").strip()
        if not value and default:
            value = default
        if value in opts:
            return value
        print(f"Invalid choice. Options: {', '.join(opts)}")


def _prompt_text(prompt: str, default: str | None = None) -> str:
    if default:
        prompt = f"{prompt} [{default}]"
    value = input(f"{prompt}: ").strip()
    return value or (default or "")


def interactive_menu(presets: list[dict], default_choice: str | None) -> dict:
    input_raw = _prompt_text("Input file or directory (MD, or multiple separated by commas)")
    inputs = [item.strip() for item in input_raw.split(",") if item.strip()]
    output_dir = ""

    print("1) Generate all outputs")
    print("2) Generate test forms only (skip copyright)")
    print("3) Generate copyright only")
    print("4) Exit")
    choice = _prompt_choice("Select an option", ["1", "2", "3", "4"], default="1")
    if choice == "4":
        return {"exit": True}

    preset_choice = None
    if presets:
        labels = [item.get("label", f"Preset {idx + 1}") for idx, item in enumerate(presets)]
        print("Contact Presets:")
        for idx, label in enumerate(labels, start=1):
            print(f"{idx}) {label}")
        selected = _prompt_choice("Select contact preset", [str(i) for i in range(1, len(labels) + 1)], default="1")
        preset_choice = labels[int(selected) - 1]
    elif default_choice:
        preset_choice = default_choice

    if choice == "2":
        skip_copyright = True
    elif choice == "3":
        skip_copyright = False
    else:
        skip_copyright = False

    applicant_type = None
    if choice != "2":
        applicant_type = _prompt_choice(
            "Applicant type (holder=personal, agent=proxy)",
            ["holder", "agent"],
            default="holder",
        )

    return {
        "inputs": inputs,
        "output_dir": output_dir,
        "skip_copyright": skip_copyright,
        "copyright_only": choice == "3",
        "preset_choice": preset_choice,
        "applicant_type": applicant_type,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate software test and copyright forms from a manual.")
    parser.add_argument("--input", "-i", nargs="*", help="Input MD file(s) or directories")
    parser.add_argument("--output-dir", "-o", default=None, help="Output directory")
    parser.add_argument("--env-file", default=str(DEFAULT_ENV_FILE), help="Env file path")
    parser.add_argument("--api-key", default="", help="LLM API key (overrides env)")
    parser.add_argument("--base-url", default="", help="LLM base URL (overrides env)")
    parser.add_argument("--model", default="", help="LLM model (overrides env)")
    parser.add_argument("--contact-info", default=str(DEFAULT_CONFIG_FILE), help="Config file (yaml or json)")
    parser.add_argument("--preset-choice", default="", help="Preset label to use")
    parser.add_argument("--applicant-type", default="", help="Applicant type: holder or agent")
    parser.add_argument("--skip-copyright", action="store_true", help="Skip copyright application output")
    parser.add_argument("--app-name", default="", help="Software name (manual input)")
    parser.add_argument("--app-version", default="", help="Software version (manual input)")
    args = parser.parse_args()

    print("DocCollate - Software form generator")

    if not ensure_local_run():
        return

    if args.env_file:
        load_dotenv(args.env_file)

    api_key = args.api_key or os.getenv("OPENAI_API_KEY", "") or os.getenv("DASHSCOPE_API_KEY", "")
    base_url = args.base_url or os.getenv("OPENAI_BASE_URL", "")
    model = args.model or os.getenv("OPENAI_MODEL", "gpt-4o-mini")

    env_config_path = os.getenv("DOCCOLLATE_CONFIG_PATH", "").strip()
    if env_config_path and args.contact_info == str(DEFAULT_CONFIG_FILE):
        args.contact_info = env_config_path

    template_paths = load_template_paths()
    if not template_paths:
        return

    if not args.app_name:
        args.app_name = _prompt_text("Software name (manual input)")
    if not args.app_version:
        args.app_version = _prompt_text("Software version (manual input)")

    default_output_base = str(Path.cwd())
    output_base_input = _prompt_text("Output base directory", default=default_output_base)
    output_base_path = Path(output_base_input).expanduser()
    if not output_base_path.exists() or not output_base_path.is_dir():
        print(f"Output base directory not found: {output_base_path}")
        return

    interactive = False
    config_path = Path(args.contact_info)
    config = load_yaml_config(config_path)
    if not args.input:
        interactive = True
        presets = config.get("presets", []) if isinstance(config, dict) else []
        default_choice = config.get("preset_choice") if isinstance(config, dict) else None
        selection = interactive_menu(presets, default_choice)
        if selection.get("exit"):
            return
        args.input = selection["inputs"]
        args.output_dir = selection["output_dir"]
        args.skip_copyright = selection["skip_copyright"]
        args.preset_choice = selection["preset_choice"] or args.preset_choice
        if selection["applicant_type"]:
            args.applicant_type = selection["applicant_type"]
        copyright_only = selection["copyright_only"]
    else:
        copyright_only = False

    output_dir = output_base_path / (args.output_dir or "")
    ensure_output_dir(output_dir)

    files = collect_inputs(args.input)
    if not files:
        print("No input files found.")
        return

    client = init_client(api_key, base_url or None)
    preset = select_preset(config if isinstance(config, dict) else {}, args.preset_choice or None)
    contact_info = resolve_contact_info(config if isinstance(config, dict) else {}, preset)
    company_profile = preset if preset else (config if isinstance(config, dict) else {})

    total = len(files)
    for idx, file_path in enumerate(files, start=1):
        print(f"Processing ({idx}/{total}): {file_path.name}")
        text = read_file_content(file_path)
        if not text:
            continue

        base_name = file_path.stem

        data: dict[str, object] = {}
        rule_data = extract_rule_data(text)
        data = merge_missing(data, rule_data)

        section_map = build_section_map(text)
        section_chunks = build_section_chunks(section_map)
        full_chunks = split_into_chunks(text)

        func_items = extract_func_items(text, section_chunks, client, model)
        if func_items:
            data["product__func_list"] = func_items
        field_data = extract_fields_by_prompt(client, model, text, section_chunks, full_chunks, FIELD_PROMPTS)

        data = merge_missing(data, field_data)

        data = sanitize_data(data)
        data = normalize_assessment_data(data)
        data = derive_fields(data)

        if args.applicant_type:
            data["applicant__type"] = args.applicant_type
        if args.app_name:
            data["app__name"] = args.app_name
        if args.app_version:
            data["app__version"] = args.app_version

        software_name = data.get("app__name") or base_name
        version = data.get("app__version") or "\u672a\u6807\u6ce8\u7248\u672c"

        if not copyright_only and template_paths["func"].exists():
            output_path = output_dir / build_filename("\u4ea7\u54c1\u6d4b\u8bd5\u529f\u80fd\u8868", software_name, version, ".docx")
            if not fill_func_table(template_paths["func"], output_path, data):
                print(f"No func table generated for {file_path}")
            else:
                print(f"生成文件: {output_path}")
        elif not copyright_only:
            print(f"Missing template: {template_paths['func']}")

        if not copyright_only and template_paths["reg"].exists():
            output_path = output_dir / build_filename("\u4ea7\u54c1\u6d4b\u8bd5\u767b\u8bb0\u8868", software_name, version, ".docx")
            if not fill_reg_table(template_paths["reg"], output_path, data, contact_info):
                print(f"No registration form generated for {file_path}")
            else:
                print(f"生成文件: {output_path}")
        elif not copyright_only:
            print(f"Missing template: {template_paths['reg']}")

        if not copyright_only and template_paths["env"].exists():
            output_path = output_dir / build_filename("\u975e\u5d4c\u5165\u5f0f\u8f6f\u4ef6\u73af\u5883", software_name, version, ".docx")
            if not fill_env_table(template_paths["env"], output_path, data):
                print(f"No environment form generated for {file_path}")
            else:
                print(f"生成文件: {output_path}")
        elif not copyright_only:
            print(f"Missing template: {template_paths['env']}")

        if not copyright_only and template_paths["assess"].exists():
            output_path = output_dir / build_filename("\u4ea7\u54c1\u8bc4\u4f30\u7533\u8bf7", software_name, version, ".xlsx")
            if not fill_assessment_excel(template_paths["assess"], output_path, data):
                print(f"No assessment form generated for {file_path}")
            else:
                print(f"生成文件: {output_path}")
        elif not copyright_only:
            print(f"Missing template: {template_paths['assess']}")

        if not args.skip_copyright and template_paths["copyright"].exists():
            output_path = output_dir / build_copyright_filename(software_name, version)
            generate_document(company_profile, data, template_paths["copyright"], output_path)
        elif not args.skip_copyright:
            print(f"Missing template: {template_paths['copyright']}")

    print(f"Done. Outputs are in {output_dir}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nCancelled.")
